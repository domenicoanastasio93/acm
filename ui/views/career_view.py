import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, messagebox
from database.db_manager import DatabaseManager
from utils.excel_util import get_unique_gestioni
from ui.components.multi_select import MultiSelectDropdown
from ui.components.delivery_dialog import DeliveryDialog
from ui.components.register_dialog import RegisterDialog
from utils.date_util import DateUtil
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill
import os

COL_GESTION = "Gestión"


class CareerView(ctk.CTkFrame):
    def __init__(self, master, career_name, theme_color, back_callback):
        super().__init__(master)
        self.career_name = career_name
        self.theme_color = theme_color
        self.back_callback = back_callback
        self.db = DatabaseManager()
        self.available_gestioni = get_unique_gestioni()

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1) # Content area

        # --- Header ---
        self.header = ctk.CTkFrame(self, fg_color=self.theme_color, height=60, corner_radius=0)
        self.header.grid(row=0, column=0, sticky="ew")
        
        self.back_btn = ctk.CTkButton(self.header, text="← Menú", width=80, fg_color="white", text_color="black", hover_color="#eee", command=self.back_callback)
        self.back_btn.pack(side="left", padx=20, pady=10)

        self.title_label = ctk.CTkLabel(self.header, text=f"{self.career_name}", font=("Roboto", 24, "bold"), text_color="white")
        self.title_label.pack(side="left", padx=20)

        # --- Main Content Area ---
        self.content = ctk.CTkFrame(self, fg_color="transparent")
        self.content.grid(row=1, column=0, sticky="nsew", padx=20, pady=20)
        self.content.grid_columnconfigure(0, weight=1) 
        self.content.grid_rowconfigure(1, weight=1)

        # --- Top: Filter Bar ---
        self.filter_frame = ctk.CTkFrame(self.content, fg_color="transparent")
        self.filter_frame.grid(row=0, column=0, sticky="ew", pady=(0, 20))
        
        # Grid for filters to keep them aligned
        for i in range(4):
            self.filter_frame.grid_columnconfigure(i, weight=1)

        # --- First Row: Basic Info ---
        # Name Filter
        self.name_var = ctk.StringVar()
        self.name_var.trace_add("write", self.filter_list)
        ctk.CTkLabel(self.filter_frame, text="Nombre Estudiante:").grid(row=0, column=0, padx=5, sticky="w")
        self.name_entry = ctk.CTkEntry(self.filter_frame, textvariable=self.name_var, placeholder_text="Filtrar por nombre...")
        self.name_entry.grid(row=1, column=0, padx=5, sticky="ew")

        # Number Filter
        self.number_var = ctk.StringVar()
        self.number_var.trace_add("write", self.filter_list)
        ctk.CTkLabel(self.filter_frame, text="Número:").grid(row=0, column=1, padx=5, sticky="w")
        self.number_entry = ctk.CTkEntry(self.filter_frame, textvariable=self.number_var, placeholder_text="Filtrar por N...")
        self.number_entry.grid(row=1, column=1, padx=5, sticky="ew")

        # Gestion Filter
        self.gestion_var = ctk.StringVar()
        self.gestion_var.trace_add("write", self.filter_list)
        ctk.CTkLabel(self.filter_frame, text="Gestión:").grid(row=0, column=2, padx=5, sticky="w")
        self.gestion_entry = ctk.CTkEntry(self.filter_frame, textvariable=self.gestion_var, placeholder_text="Filtrar por gestión...")
        self.gestion_entry.grid(row=1, column=2, padx=5, sticky="ew")

        # Status Filter
        self.status_list = ["Todos", "Pendientes", "Entregados"]
        self.status_var = ctk.StringVar(value="Todos")
        self.status_var.trace_add("write", self.filter_list)
        ctk.CTkLabel(self.filter_frame, text="Estado:").grid(row=0, column=3, padx=5, sticky="w")
        self.status_menu = ctk.CTkOptionMenu(self.filter_frame, values=self.status_list, variable=self.status_var, width=100)
        self.status_menu.grid(row=1, column=3, padx=5, sticky="ew")

        # --- Second Row: Date Filters ---
        self.months_list = ["Todos", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        current_year = datetime.datetime.now().year
        self.years_list = ["Todos"] + [str(y) for y in range(current_year - 5, current_year + 5)]

        # Insertion Month Filter
        self.month_var = ctk.StringVar(value="Todos")
        self.month_var.trace_add("write", self.filter_list)
        ctk.CTkLabel(self.filter_frame, text="Mes Inserción:").grid(row=2, column=0, padx=5, sticky="w", pady=(10, 0))
        self.month_menu = ctk.CTkOptionMenu(self.filter_frame, values=self.months_list, variable=self.month_var, width=100)
        self.month_menu.grid(row=3, column=0, padx=5, sticky="ew")

        # Insertion Year Filter
        self.year_var = ctk.StringVar(value="Todos")
        self.year_var.trace_add("write", self.filter_list)
        ctk.CTkLabel(self.filter_frame, text="Año Inserción:").grid(row=2, column=1, padx=5, sticky="w", pady=(10, 0))
        self.year_menu = ctk.CTkOptionMenu(self.filter_frame, values=self.years_list, variable=self.year_var, width=100)
        self.year_menu.grid(row=3, column=1, padx=5, sticky="ew")

        # Delivery Month Filter
        self.delivery_month_var = ctk.StringVar(value="Todos")
        self.delivery_month_var.trace_add("write", self.filter_list)
        ctk.CTkLabel(self.filter_frame, text="Mes Entrega:").grid(row=2, column=2, padx=5, sticky="w", pady=(10, 0))
        self.delivery_month_menu = ctk.CTkOptionMenu(self.filter_frame, values=self.months_list, variable=self.delivery_month_var, width=100)
        self.delivery_month_menu.grid(row=3, column=2, padx=5, sticky="ew")

        # Delivery Year Filter
        self.delivery_year_var = ctk.StringVar(value="Todos")
        self.delivery_year_var.trace_add("write", self.filter_list)
        ctk.CTkLabel(self.filter_frame, text="Año Entrega:").grid(row=2, column=3, padx=5, sticky="w", pady=(10, 0))
        self.delivery_year_menu = ctk.CTkOptionMenu(self.filter_frame, values=self.years_list, variable=self.delivery_year_var, width=100)
        self.delivery_year_menu.grid(row=3, column=3, padx=5, sticky="ew")

        # Total Count Label (moved to actions_frame)

        # --- Center: Student List ---
        self.list_frame = ctk.CTkFrame(self.content)
        self.list_frame.grid(row=1, column=0, sticky="nsew")
        self.list_frame.grid_rowconfigure(0, weight=1)
        self.list_frame.grid_columnconfigure(0, weight=1)

        # Treeview Style
        style = ttk.Style()
        style.theme_use("default")
        style.configure("Treeview", 
                        background="#2b2b2b", 
                        foreground="white", 
                        fieldbackground="#2b2b2b", 
                        rowheight=30)
        style.map("Treeview", background=[("selected", self.theme_color)])
        style.configure("Treeview.Heading", font=("Roboto", 12, "bold"))
        
        # Treeview
        # We keep ID as the first column but hide it from display, and add delivery date.
        columns = ("db_id", "Número", "Nombre", "Factura", COL_GESTION, "Fecha", "Entrega", "Notas")
        self.tree = ttk.Treeview(self.list_frame, columns=columns, show="headings", selectmode="browse")
        
        # Hide the db_id column from display
        self.tree["displaycolumns"] = ("Número", "Nombre", "Factura", COL_GESTION, "Fecha", "Entrega", "Notas")

        self.tree.heading("Número", text="Número")
        self.tree.heading("Nombre", text="Nombre")
        self.tree.heading("Factura", text="Factura")
        self.tree.heading(COL_GESTION, text=COL_GESTION)
        self.tree.heading("Fecha", text="Inserción")
        self.tree.heading("Entrega", text="Entrega")
        self.tree.heading("Notas", text="Note")

        self.tree.column("Número", width=50)
        self.tree.column("Nombre", width=200)
        self.tree.column("Factura", width=80)
        self.tree.column(COL_GESTION, width=150)
        self.tree.column("Fecha", width=120)
        self.tree.column("Entrega", width=120)
        self.tree.column("Notas", width=150)


        # Treeview
        self.tree.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)

        # Scrollbar for tree
        scrollbar = ttk.Scrollbar(self.list_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.grid(row=0, column=1, sticky="ns", pady=10)

        # Action Buttons (Bottom Center)
        self.actions_frame = ctk.CTkFrame(self.content, fg_color="transparent")
        self.actions_frame.grid(row=2, column=0, sticky="ew", pady=(10, 0))
        
        # Grid layout for actions_frame to keep buttons centered and label on the right
        self.actions_frame.grid_columnconfigure(0, weight=1) # Left spacer
        self.actions_frame.grid_columnconfigure(1, weight=0) # Buttons area
        self.actions_frame.grid_columnconfigure(2, weight=1) # Right area for label
        
        # Internal container to keep buttons centered
        self.buttons_inner = ctk.CTkFrame(self.actions_frame, fg_color="transparent")
        self.buttons_inner.grid(row=0, column=1)
        
        self.btn_new = ctk.CTkButton(self.buttons_inner, text="+ Nuevo Registro", width=200, height=40, fg_color="#27ae60", hover_color="#219150", font=("Roboto", 13, "bold"), command=self.open_register_modal)
        self.btn_new.grid(row=0, column=0, padx=10)

        self.btn_deliver = ctk.CTkButton(self.buttons_inner, text="Entregar Certificado", width=200, height=40, fg_color="gray", state="disabled", command=self.deliver_certificate)
        self.btn_deliver.grid(row=0, column=1, padx=10)

        self.btn_delete = ctk.CTkButton(self.buttons_inner, text="Eliminar Registro", width=200, height=40, fg_color="gray", state="disabled", hover_color="#c0392b", command=self.delete_record)
        self.btn_delete.grid(row=0, column=2, padx=10)

        self.btn_export = ctk.CTkButton(self.buttons_inner, text="Exportar Excel", width=200, height=40, fg_color="#2196F3", hover_color="#1976D2", font=("Roboto", 13, "bold"), command=self.do_export)
        self.btn_export.grid(row=0, column=3, padx=10)

        # Total Count Label - Placed in column 2 (sticky right)
        self.count_label = ctk.CTkLabel(self.actions_frame, text="Total Certificados: 0", font=("Roboto", 14, "bold"))
        self.count_label.grid(row=0, column=2, sticky="e", padx=(10, 20))

        self.tree.bind("<<TreeviewSelect>>", self.on_select)
        self.tree.bind("<Double-1>", self.on_double_click)
        
        # Context menu (Right click)
        self.menu = tk.Menu(self, tearoff=0)
        self.menu.add_command(label="Entregar Certificado", command=self.deliver_certificate)
        self.menu.add_command(label="Eliminar Registro", command=self.delete_record)
        self.tree.bind("<Button-2>" if self.tk.call('tk', 'windowingsystem') == 'aqua' else "<Button-3>", self.show_menu)
        
        # Tag configuration for delivered items
        self.tree.tag_configure("delivered", foreground="#00E676")

        self.refresh_data()

    def open_register_modal(self):
        RegisterDialog(self.master, self.available_gestioni, self.theme_color, self.add_record)

    def add_record(self, numero, name, factura, gestion_str, notas=None):
        if name:
            self.db.add_certificate(numero, name, self.career_name, factura, gestion_str, notas)
            self.refresh_data()
        else:
            messagebox.showwarning("Atención", "El nombre del estudiante es obligatorio.")

    def refresh_data(self):
        # We just call filter_list as it already handles clearing, fetching and displaying
        self.filter_list()

    def filter_list(self, *args):
        name_query = self.name_var.get().lower()
        num_query = self.number_var.get().lower()
        gestion_query = self.gestion_var.get().lower()
        
        selected_month_name = self.month_var.get()
        selected_month = self.months_list.index(selected_month_name) if selected_month_name != "Todos" else None
        selected_year = self.year_var.get()
        year_filter = selected_year if selected_year != "Todos" else None

        sel_del_month_name = self.delivery_month_var.get()
        sel_del_month = self.months_list.index(sel_del_month_name) if sel_del_month_name != "Todos" else None
        sel_del_year = self.delivery_year_var.get()
        del_year_filter = sel_del_year if sel_del_year != "Todos" else None

        selected_status = self.status_var.get()

        for item in self.tree.get_children():
            self.tree.delete(item)
            
        records = self.db.get_career_certificates(self.career_name)
        count = 0
        for r in records:
            # r: (id, numero, nombre, factura, gestion, fecha_registro, estado, notas, fecha_entrega)
            numero = str(r[1]) if r[1] else ""
            name = str(r[2]) if r[2] else ""
            factura = str(r[3]) if r[3] else ""
            gestion_str = r[4] if r[4] else ""
            fecha_registro_raw = r[5] # YYYY-MM-DD HH:MM:SS
            fecha_formattata = DateUtil.format_datetime(fecha_registro_raw)
            estado = r[6]
            notas = r[7]
            
            # Date filter logic
            date_match = True
            if fecha_registro_raw:
                try:
                    dt = datetime.datetime.strptime(fecha_registro_raw, "%Y-%m-%d %H:%M:%S")
                    if selected_month and dt.month != selected_month:
                        date_match = False
                    if year_filter and str(dt.year) != year_filter:
                        date_match = False
                except ValueError:
                    pass

            # Date match for delivery
            delivery_match = True
            fecha_entrega_raw = r[8] # YYYY-MM-DD HH:MM:SS
            if sel_del_month or del_year_filter:
                if not fecha_entrega_raw:
                    delivery_match = False
                else:
                    try:
                        dt_del = datetime.datetime.strptime(fecha_entrega_raw, "%Y-%m-%d %H:%M:%S")
                        if sel_del_month and dt_del.month != sel_del_month:
                            delivery_match = False
                        if del_year_filter and str(dt_del.year) != del_year_filter:
                            delivery_match = False
                    except ValueError:
                        delivery_match = False
            
            # Status filter logic
            status_match = True
            if selected_status == "Pendientes":
                if estado != 0:
                    status_match = False
            elif selected_status == "Entregados":
                if estado != 1:
                    status_match = False

            # Apply individual filters
            if (name_query in name.lower() and 
                (num_query in numero.lower() or num_query in factura.lower()) and
                gestion_query in gestion_str.lower() and
                date_match and delivery_match and status_match):
                
                count += self._count_certificates_in_record(gestion_str)
                
                # Create the record for the treeview
                # columns = ("db_id", "N.", "Nombre", "Factura", "Gestión", "Fecha", "Entrega", "Notas")
                fecha_entrega = r[8]
                fecha_entrega_formattata = DateUtil.format_datetime(fecha_entrega)
                formatted_record = [r[0], numero, name, factura, gestion_str, fecha_formattata, fecha_entrega_formattata, notas]
                
                tags = ("delivered",) if estado == 1 else ()
                self.tree.insert("", "end", values=formatted_record, tags=tags)
        
        self.count_label.configure(text=f"Total Certificados: {count}")

    def _count_certificates_in_record(self, gestion_str):
        """
        Calcula el número de certificados físicos usando una única búsqueda unificada.
        Ejemplo: "I-24 (X3) II-24" -> 3 + 1 = 4 total.
        """
        import re
        if not gestion_str or not str(gestion_str).strip():
            return 0
            
        # Busca el patrón: SEMESTRE seguido opcionalmente por (Xn)
        # Grupo 1: El semestre (es. I-24)
        # Grupo 2: La cantidad (es. 3)
        pattern = r'((?:II|I|VER|INV)-\d+)(?:\s*\(?[xX](\d+)\)?)?'
        matches = re.findall(pattern, gestion_str)
        
        # Se non trova pattern standard, ma c'è del testo, lo conta come 1
        if not matches:
            return 1 if gestion_str.strip() else 0
            
        total = 0
        for _, multiplier in matches:
            # Se c'è un moltiplicatore, usa quello, altrimenti conta 1
            total += int(multiplier) if multiplier else 1
            
        return total

    def on_select(self, event):
        selected = self.tree.selection()
        if selected:
            tags = self.tree.item(selected[0])['tags']
            is_delivered = "delivered" in tags
            
            if is_delivered:
                self.btn_deliver.configure(text="Anular Entrega", state="normal", fg_color="#f39c12", hover_color="#d3840e", command=self.undo_delivery)
            else:
                self.btn_deliver.configure(text="Entregar Certificado", state="normal", fg_color="#9512f3", hover_color="#7d0fca", command=self.deliver_certificate)
                 
            self.btn_delete.configure(state="normal", fg_color="#e74c3c")
        else:
            self.btn_deliver.configure(text="Entregar Certificado", state="disabled", fg_color="gray")
            self.btn_delete.configure(state="disabled", fg_color="gray")

    def show_menu(self, event):
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)
            # Dynamic menu based on status
            tags = self.tree.item(item)['tags']
            is_delivered = "delivered" in tags
            
            self.menu.delete(0, "end")
            if is_delivered:
                self.menu.add_command(label="Anular Entrega", command=self.undo_delivery)
            else:
                self.menu.add_command(label="Entregar Certificado", command=self.deliver_certificate)
            self.menu.add_command(label="Eliminar Registro", command=self.delete_record)
            
            self.menu.post(event.x_root, event.y_root)

    def deliver_certificate(self):
        selected = self.tree.selection()
        if not selected: return
        
        item_values = self.tree.item(selected[0])['values']
        cert_id, name = item_values[0], item_values[2] # 0 is db_id, 2 is name

        def on_confirm(custom_date):
            self.db.mark_as_delivered(cert_id, custom_date)
            self.refresh_data()
            self.btn_deliver.configure(text="Entregar Certificado", state="disabled", fg_color="gray")
            self.btn_delete.configure(state="disabled", fg_color="gray")

        DeliveryDialog(self.master, name, self.theme_color, on_confirm)

    def undo_delivery(self):
        selected = self.tree.selection()
        if not selected: return
        
        item_values = self.tree.item(selected[0])['values']
        cert_id, name = item_values[0], item_values[2] # 0 is db_id, 2 is name

        if messagebox.askyesno("Confirmar Restablecimiento", f"¿Desea anular la entrega para {name}?"):
            self.db.undo_delivery(cert_id)
            self.refresh_data()
            self.btn_deliver.configure(text="Entregar Certificado", state="disabled", fg_color="gray")
            self.btn_delete.configure(state="disabled", fg_color="gray")

    def delete_record(self):
        selected = self.tree.selection()
        if not selected: return
        
        item_values = self.tree.item(selected[0])['values']
        cert_id, name = item_values[0], item_values[2] # 0 is db_id, 2 is name

        if messagebox.askyesno("Confirmar Eliminación", f"¿Está seguro de que desea eliminar a {name}?\nEsta operación es irreversible."):
             self.db.delete_certificate(cert_id)
             self.refresh_data()
             self.btn_deliver.configure(text="Entregar Certificado", state="disabled", fg_color="gray")
             self.btn_delete.configure(state="disabled", fg_color="gray")

    def do_export(self):
        try:
            # Get data from treeview
            items = self.tree.get_children()
            if not items:
                messagebox.showinfo("Información", "No hay datos para exportar.")
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Report {self.career_name}"

            headers = ["Número", "Nombre Estudiante", "Factura", COL_GESTION, "Fecha Registro", "Fecha Entrega", "Notas"]
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)

            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Standard Excel Light Green

            for item in items:
                values = self.tree.item(item)['values']
                tags = self.tree.item(item)['tags']
                # values: [db_id, Número, Nombre, Factura, Gestión, Fecha, Entrega, Notas]
                # Skip db_id (index 0)
                ws.append(values[1:])
                
                if "delivered" in tags:
                    row_idx = ws.max_row
                    for cell in ws[row_idx]:
                        cell.fill = green_fill

            now = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"Report_{self.career_name}_{now}.xlsx"
            
            # Use Downloads folder if it exists, otherwise use Home
            home = os.path.expanduser("~")
            downloads = os.path.join(home, "Downloads")
            if os.path.exists(downloads):
                save_path = os.path.join(downloads, filename)
            else:
                save_path = os.path.join(home, filename)
            
            wb.save(save_path)
            messagebox.showinfo("Éxito", f"Archivo guardado: {save_path}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error durante la exportación: {e}")

    def on_double_click(self, event):
        """Handle double-click to edit a cell."""
        region = self.tree.identify_region(event.x, event.y)
        if region != "cell":
            return

        column = self.tree.identify_column(event.x) # e.g. "#1"
        item = self.tree.identify_row(event.y)
        
        # Get column index (1-based because of how displaycolumns/columns work here)
        try:
            col_idx = int(column[1:])
        except (ValueError, IndexError):
            return
        
        # Display Columns: ("Número", "Nombre", "Factura", COL_GESTION, "Fecha", "Entrega", "Notas")
        # Indices:           1      2         3          4           5        6          7
        
        # Non-editable columns: 5 (Fecha/Insercion) and 6 (Entrega)
        if col_idx in [5, 6]:
            return

        # Get column name from display columns
        display_cols = self.tree["displaycolumns"]
        col_name = display_cols[col_idx - 1]

        # Get the current value
        current_values = list(self.tree.item(item, 'values'))
        db_id = current_values[0]
        
        # Mapping display column to the index in the full 'values' list
        # values: [id, numero, nombre, factura, gestion, fecha, entrega, notas]
        # index:  0   1       2       3        4        5      6        7
        val_idx = col_idx # conveniently maps 1:1 for most, but let's be careful
        
        old_value = current_values[val_idx]

        # Get cell coordinates
        try:
            bbox = self.tree.bbox(item, column)
            if not bbox: return
            x, y, width, height = bbox
        except Exception:
            return

        # Create entry widget with matching theme colors
        entry = tk.Entry(self.tree, 
                         font=("Roboto", 11), 
                         bg="#1e1e1e", 
                         fg="white", 
                         insertbackground="white",
                         relief="flat",
                         borderwidth=1)
        entry.insert(0, str(old_value))
        entry.place(x=x, y=y, width=width, height=height)
        entry.focus_set()
        
        # Select all text
        entry.after(10, lambda: entry.select_range(0, 'end'))

        def save_edit(event=None):
            if not entry.winfo_exists():
                return
            new_value = entry.get()
            if str(new_value) != str(old_value):
                # Update DB
                self.db.update_certificate_field(db_id, col_name, new_value)
                # Update Treeview
                current_values[val_idx] = new_value
                self.tree.item(item, values=current_values)
                # If Gestión changed, refresh to update count
                if col_name == COL_GESTION:
                    self.refresh_data()
            entry.destroy()

        def cancel_edit(event=None):
            if entry.winfo_exists():
                entry.destroy()

        entry.bind("<Return>", save_edit)
        entry.bind("<Escape>", cancel_edit)
        entry.bind("<FocusOut>", lambda e: save_edit())
